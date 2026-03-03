from __future__ import annotations

import calendar
import json
import os
import uuid
from dataclasses import dataclass, asdict
from datetime import date
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import ttkbootstrap as tb
from ttkbootstrap.constants import *
from ttkbootstrap.dialogs import Messagebox

# ✅ import
from ttkbootstrap.widgets.tableview import Tableview

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# =========================
# Constantes
# =========================
SHIFT_ORDER = ["Manhã", "Intermediário", "Tarde", "Noite", "Madrugada"]
ROLES = ["Jr", "Pl", "Sr"]

MONTHS_PT = [
    "", "Janeiro", "Fevereiro", "Março", "Abril", "Maio", "Junho",
    "Julho", "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro"
]

PT_WEEKDAY_SUN_FIRST = ["Dom", "Seg", "Ter", "Qua", "Qui", "Sex", "Sab"]

CONFIG_FILE = "escala_config.json"

SHIFT_COLOR = {
    "Manhã": "6AA84F",
    "Intermediário": "F1C232",
    "Tarde": "3C78D8",
    "Noite": "8E7CC3",
    "Madrugada": "674EA7",
}


# =========================
# Modelo
# =========================
@dataclass
class Employee:
    emp_id: str
    team: str
    name: str
    role: str
    shift: str
    hours: str
    off_days: List[int]
    vacation_start: Optional[int] = None
    vacation_end: Optional[int] = None


def new_id() -> str:
    return uuid.uuid4().hex


def safe_int_list_from_csv(text: str) -> List[int]:
    text = (text or "").strip()
    if not text:
        return []
    out: List[int] = []
    for part in text.replace(";", ",").split(","):
        p = part.strip()
        if p.isdigit():
            out.append(int(p))
    seen = set()
    uniq: List[int] = []
    for x in out:
        if x not in seen:
            uniq.append(x)
            seen.add(x)
    return uniq


def month_days(year: int, month: int) -> List[date]:
    _, last_day = calendar.monthrange(year, month)
    return [date(year, month, d) for d in range(1, last_day + 1)]


def is_weekend(d: date) -> bool:
    return d.weekday() >= 5


def build_headers_for_month(days: List[date]) -> List[str]:
    headers = []
    for d in days:
        dow = PT_WEEKDAY_SUN_FIRST[(d.weekday() + 1) % 7]
        headers.append(f"{d.day} - {dow}")
    return headers


# =========================
# Persistência
# =========================
def load_config() -> Tuple[List[Employee], int, int]:
    today = date.today()
    if not os.path.exists(CONFIG_FILE):
        return [], today.year, today.month

    try:
        with open(CONFIG_FILE, "r", encoding="utf-8") as f:
            data = json.load(f)

        emps: List[Employee] = []
        for x in data.get("employees", []):
            emps.append(
                Employee(
                    emp_id=x.get("emp_id") or new_id(),
                    team=x.get("team", ""),
                    name=x.get("name", ""),
                    role=x.get("role", "Jr"),
                    shift=x.get("shift", SHIFT_ORDER[0]),
                    hours=x.get("hours", "-"),
                    off_days=x.get("off_days", []) or [],
                    vacation_start=x.get("vacation_start", None),
                    vacation_end=x.get("vacation_end", None),
                )
            )

        year = int(data.get("year", today.year))
        month = int(data.get("month", today.month))
        return emps, year, month
    except Exception:
        return [], today.year, today.month


def save_config(employees: List[Employee], year: int, month: int) -> None:
    data = {"year": year, "month": month, "employees": [asdict(e) for e in employees]}
    with open(CONFIG_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


# =========================
# Excel (Cambria 11, P/F/Férias, borda externa grossa por bloco, freeze panes)
# =========================
def set_outer_border(ws, r1: int, c1: int, r2: int, c2: int, side: Side):
    for c in range(c1, c2 + 1):
        top_cell = ws.cell(row=r1, column=c)
        bottom_cell = ws.cell(row=r2, column=c)
        top_cell.border = Border(
            left=top_cell.border.left, right=top_cell.border.right,
            top=side, bottom=top_cell.border.bottom
        )
        bottom_cell.border = Border(
            left=bottom_cell.border.left, right=bottom_cell.border.right,
            top=bottom_cell.border.top, bottom=side
        )

    for r in range(r1, r2 + 1):
        left_cell = ws.cell(row=r, column=c1)
        right_cell = ws.cell(row=r, column=c2)
        left_cell.border = Border(
            left=side, right=left_cell.border.right,
            top=left_cell.border.top, bottom=left_cell.border.bottom
        )
        right_cell.border = Border(
            left=right_cell.border.left, right=side,
            top=right_cell.border.top, bottom=right_cell.border.bottom
        )


def generate_schedule_xlsx(employees: List[Employee], year: int, month: int, output_path: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = f"{MONTHS_PT[month]} {year}"

    days = month_days(year, month)
    day_headers = build_headers_for_month(days)

    fixed_cols = ["Time", "Nome", "Cargo", "Turno", "Horário"]
    fixed_count = len(fixed_cols)
    total_cols = fixed_count + len(day_headers)

    base_font = Font(name="Cambria", size=11)
    base_bold = Font(name="Cambria", size=11, bold=True)

    thin = Side(style="thin", color="666666")
    thick = Side(style="thick", color="333333")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    fill_header = PatternFill("solid", fgColor="D9D9D9")
    fill_weekend_hdr = PatternFill("solid", fgColor="CFE2F3")
    fill_off = PatternFill("solid", fgColor="FF4D4D")
    fill_present = PatternFill("solid", fgColor="93C47D")
    fill_team_sep = PatternFill("solid", fgColor="F2F2F2")
    fill_vacation = PatternFill("solid", fgColor="FFD966")

    font_section = Font(name="Cambria", size=11, bold=True, color="FFFFFF")
    font_off = Font(name="Cambria", size=11, bold=True, color="FFFFFF")
    font_present = Font(name="Cambria", size=11, bold=True, color="0B3D0B")
    font_vacation = Font(name="Cambria", size=11, bold=True, color="7F6000")

    by_shift: Dict[str, List[Employee]] = {s: [] for s in SHIFT_ORDER}
    for e in employees:
        by_shift[e.shift if e.shift in by_shift else SHIFT_ORDER[0]].append(e)
    for s in SHIFT_ORDER:
        by_shift[s].sort(key=lambda x: (x.team.lower(), x.name.lower()))

    def style_cell(cell, font=None, fill=None, align=None):
        cell.font = font or base_font
        if fill is not None:
            cell.fill = fill
        cell.alignment = align or align_center
        cell.border = border_thin

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 26
    ws.column_dimensions["C"].width = 8
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 18
    for i in range(len(day_headers)):
        ws.column_dimensions[get_column_letter(fixed_count + 1 + i)].width = 9

    row = 1
    first_block_data_start_row: Optional[int] = None

    for shift in SHIFT_ORDER:
        group = by_shift.get(shift, [])
        if not group:
            continue

        block_start = row

        fill_shift = PatternFill("solid", fgColor=SHIFT_COLOR.get(shift, "333333"))
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=total_cols)
        cell = ws.cell(row=row, column=1, value=f"Turno: {shift}")
        style_cell(cell, font=font_section, fill=fill_shift, align=align_left)
        ws.row_dimensions[row].height = 22
        row += 1

        for c in range(1, fixed_count + 1):
            cell = ws.cell(row=row, column=c, value="")
            style_cell(cell, font=base_bold, fill=fill_header, align=align_center)
        ws.merge_cells(start_row=row, start_column=fixed_count + 1, end_row=row, end_column=total_cols)
        cell = ws.cell(row=row, column=fixed_count + 1, value=MONTHS_PT[month].upper())
        style_cell(cell, font=base_bold, fill=fill_header, align=align_center)
        ws.row_dimensions[row].height = 20
        row += 1

        headers = fixed_cols + day_headers
        for i, h in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=i, value=h)
            style_cell(cell, font=base_bold, fill=fill_header, align=align_center)
        for i, d in enumerate(days):
            if is_weekend(d):
                ws.cell(row=row, column=fixed_count + 1 + i).fill = fill_weekend_hdr
        ws.row_dimensions[row].height = 34
        row += 1

        if first_block_data_start_row is None:
            first_block_data_start_row = row

        last_team = None
        for e in group:
            if e.team != last_team:
                ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=total_cols)
                cell = ws.cell(row=row, column=1, value=f"Time: {e.team}")
                style_cell(cell, font=base_bold, fill=fill_team_sep, align=align_left)
                ws.row_dimensions[row].height = 18
                row += 1
                last_team = e.team

            fixed_vals = [e.team, e.name, e.role, e.shift, e.hours]
            for c, v in enumerate(fixed_vals, start=1):
                cell = ws.cell(row=row, column=c, value=v)
                style_cell(cell, font=base_font, fill=None,
                           align=align_left if c in (1, 2, 5) else align_center)

            off_set = set(e.off_days or [])
            vac_set = set()
            if isinstance(e.vacation_start, int) and isinstance(e.vacation_end, int) and e.vacation_start <= e.vacation_end:
                vac_set = set(range(e.vacation_start, e.vacation_end + 1))

            for i, d in enumerate(days):
                col = fixed_count + 1 + i
                cell = ws.cell(row=row, column=col)
                style_cell(cell, font=base_font, fill=None, align=align_center)

                if d.day in vac_set:
                    cell.value = "Férias"
                    cell.fill = fill_vacation
                    cell.font = font_vacation
                elif d.day in off_set:
                    cell.value = "F"
                    cell.fill = fill_off
                    cell.font = font_off
                else:
                    cell.value = "P"
                    cell.fill = fill_present
                    cell.font = font_present

            ws.row_dimensions[row].height = 18
            row += 1

        row += 1
        block_end = row - 1
        set_outer_border(ws, block_start, 1, block_end, total_cols, thick)

    if first_block_data_start_row is not None:
        ws.freeze_panes = f"F{first_block_data_start_row}"

    wb.save(output_path)


# =========================
# App Tk (ttkbootstrap)
# =========================
class App(tb.Window):
    def __init__(self):
        super().__init__(themename="darkly")
        self.title("Gerador de Escala (Excel)")
        self.geometry("1280x820")

        self.employees, self.year, self.month = load_config()
        self.last_generated_path: Optional[str] = None
        self.emp_by_id: Dict[str, Employee] = {e.emp_id: e for e in self.employees}

        top = tb.Frame(self, padding=10)
        top.pack(fill=X)

        tb.Label(top, text="Ano").pack(side=LEFT)
        self.year_var = tb.StringVar(value=str(self.year))
        tb.Combobox(
            top, textvariable=self.year_var, width=6,
            values=[str(y) for y in range(date.today().year - 2, date.today().year + 6)]
        ).pack(side=LEFT, padx=8)

        tb.Label(top, text="Mês").pack(side=LEFT)
        self.month_var = tb.StringVar(value=str(self.month))
        tb.Combobox(
            top, textvariable=self.month_var, width=16,
            values=[str(i) for i in range(1, 13)]
        ).pack(side=LEFT, padx=8)
        tb.Label(top, text="(1=Jan ... 12=Dez)").pack(side=LEFT, padx=6)

        tb.Button(top, text="Adicionar colaborador", bootstyle=SUCCESS, command=self.open_add_dialog).pack(side=LEFT, padx=10)
        tb.Button(top, text="Editar selecionado", bootstyle=WARNING, command=self.open_edit_dialog).pack(side=LEFT, padx=6)
        tb.Button(top, text="Excluir selecionado", bootstyle=DANGER, command=self.delete_selected).pack(side=LEFT, padx=6)

        tb.Button(top, text="Gerar Excel", bootstyle=PRIMARY, command=self.generate_excel).pack(side=LEFT, padx=16)
        tb.Button(top, text="Abrir pasta", bootstyle=SECONDARY, command=self.open_folder).pack(side=LEFT, padx=6)

        body = tb.Frame(self, padding=10)
        body.pack(fill=BOTH, expand=True)

        left = tb.Labelframe(body, text="Colaboradores", padding=10)
        left.pack(side=LEFT, fill=BOTH, expand=True, padx=(0, 10))

        right = tb.Labelframe(body, text="Preview (todos os turnos)", padding=10)
        right.pack(side=LEFT, fill=BOTH, expand=True)

        # ✅ ID escondido como primeira coluna -> Editar/Excluir sempre com Search
        self.emp_table = Tableview(
            master=left,
            coldata=[
                {"text": "ID", "stretch": False, "width": 0},
                {"text": "Time", "stretch": True},
                {"text": "Nome", "stretch": True},
                {"text": "Cargo", "stretch": False, "width": 70},
                {"text": "Turno", "stretch": True},
                {"text": "Horário", "stretch": True},
                {"text": "Folgas", "stretch": True},
                {"text": "Férias", "stretch": True},
            ],
            rowdata=[],
            paginated=False,
            searchable=True,
            autofit=True,
            height=22,
            bootstyle=PRIMARY,
        )
        self.emp_table.pack(fill=BOTH, expand=True)

        self.preview_canvas = tb.Canvas(right, highlightthickness=0)
        self.preview_scroll = tb.Scrollbar(right, orient=VERTICAL, command=self.preview_canvas.yview)
        self.preview_canvas.configure(yscrollcommand=self.preview_scroll.set)
        self.preview_scroll.pack(side=RIGHT, fill=Y)
        self.preview_canvas.pack(side=LEFT, fill=BOTH, expand=True)

        self.preview_inner = tb.Frame(self.preview_canvas)
        self.preview_canvas.create_window((0, 0), window=self.preview_inner, anchor="nw")
        self.preview_inner.bind("<Configure>", lambda e: self.preview_canvas.configure(scrollregion=self.preview_canvas.bbox("all")))

        self.status_var = tb.StringVar(value="P = Presente (verde) | F = Folga (vermelho) | Férias = amarelo")
        tb.Label(self, textvariable=self.status_var, padding=10).pack(fill=X)

        self.refresh_employees_table()
        self.refresh_preview_all()

    def get_year_month(self) -> Tuple[int, int]:
        try:
            y = int(self.year_var.get())
            m = int(self.month_var.get())
            if m < 1 or m > 12:
                raise ValueError
            return y, m
        except Exception:
            Messagebox.show_error("Ano/Mês inválido", "Preencha um ano válido e o mês de 1 a 12.")
            return self.year, self.month

    def persist(self):
        y, m = self.get_year_month()
        save_config(self.employees, y, m)

    def validate_day(self, val: Optional[str], y: int, m: int) -> Optional[int]:
        s = (val or "").strip()
        if not s:
            return None
        if not s.isdigit():
            return None
        d = int(s)
        _, last = calendar.monthrange(y, m)
        return d if 1 <= d <= last else None

    def validate_off_days(self, off_days: List[int], y: int, m: int) -> List[int]:
        _, last = calendar.monthrange(y, m)
        return [d for d in off_days if 1 <= d <= last]

    def refresh_employees_table(self):
        self.emp_table.delete_rows()
        self.emp_by_id = {e.emp_id: e for e in self.employees}

        rows = []
        for e in self.employees:
            vac_txt = ""
            if isinstance(e.vacation_start, int) and isinstance(e.vacation_end, int):
                vac_txt = f"{e.vacation_start}-{e.vacation_end}"

            rows.append([
                e.emp_id,  # ✅ chave real
                e.team,
                e.name,
                e.role,
                e.shift,
                e.hours,
                ",".join(map(str, e.off_days or [])),
                vac_txt,
            ])

        self.emp_table.insert_rows("end", rows)

    def selected_employee(self) -> Optional[Employee]:
        tree = self.emp_table.view
        sel = tree.selection()
        if not sel:
            return None
        values = tree.item(sel[0], "values")
        if not values:
            return None
        emp_id = values[0]
        return self.emp_by_id.get(emp_id)

    def open_add_dialog(self):
        self.open_employee_dialog(None)

    def open_edit_dialog(self):
        emp = self.selected_employee()
        if not emp:
            Messagebox.show_warning("Seleção", "Selecione um colaborador para editar.")
            return
        self.open_employee_dialog(emp.emp_id)

    def delete_selected(self):
        emp = self.selected_employee()
        if not emp:
            Messagebox.show_warning("Seleção", "Selecione um colaborador para excluir.")
            return

        ok = Messagebox.okcancel("Confirmar", f"Excluir '{emp.name}'?")
        if not ok:
            return

        self.employees = [e for e in self.employees if e.emp_id != emp.emp_id]
        self.persist()
        self.refresh_employees_table()
        self.refresh_preview_all()

    def open_employee_dialog(self, emp_id: Optional[str]):
        y, m = self.get_year_month()
        existing = self.emp_by_id.get(emp_id) if emp_id else None

        win = tb.Toplevel(self)
        win.title("Colaborador")
        win.geometry("560x520")
        win.transient(self)
        win.grab_set()

        frame = tb.Frame(win, padding=16)
        frame.pack(fill=BOTH, expand=True)

        team_var = tb.StringVar(value=(existing.team if existing else ""))
        name_var = tb.StringVar(value=(existing.name if existing else ""))
        role_var = tb.StringVar(value=(existing.role if existing else "Jr"))
        shift_var = tb.StringVar(value=(existing.shift if existing else SHIFT_ORDER[0]))
        hours_var = tb.StringVar(value=(existing.hours if existing else ""))
        off_var = tb.StringVar(value=(",".join(map(str, existing.off_days)) if existing else ""))

        vac_start_var = tb.StringVar(value=("" if not existing or existing.vacation_start is None else str(existing.vacation_start)))
        vac_end_var = tb.StringVar(value=("" if not existing or existing.vacation_end is None else str(existing.vacation_end)))

        tb.Label(frame, text="TIME").pack(anchor=W)
        tb.Entry(frame, textvariable=team_var).pack(fill=X, pady=(0, 10))

        tb.Label(frame, text="Nome").pack(anchor=W)
        tb.Entry(frame, textvariable=name_var).pack(fill=X, pady=(0, 10))

        row = tb.Frame(frame)
        row.pack(fill=X, pady=(0, 10))
        tb.Label(row, text="Cargo").pack(side=LEFT)
        tb.Combobox(row, textvariable=role_var, values=ROLES, width=8).pack(side=LEFT, padx=8)
        tb.Label(row, text="Turno").pack(side=LEFT)
        tb.Combobox(row, textvariable=shift_var, values=SHIFT_ORDER, width=18).pack(side=LEFT, padx=8)

        tb.Label(frame, text="Horário (ex: 06:00 às 15:00)").pack(anchor=W)
        tb.Entry(frame, textvariable=hours_var).pack(fill=X, pady=(0, 10))

        tb.Label(frame, text="Folgas (dias do mês) ex: 7, 8, 15").pack(anchor=W)
        tb.Entry(frame, textvariable=off_var).pack(fill=X, pady=(0, 10))

        tb.Separator(frame).pack(fill=X, pady=12)

        tb.Label(frame, text="Férias (opcional)", bootstyle=INFO).pack(anchor=W)
        vac_row = tb.Frame(frame)
        vac_row.pack(fill=X, pady=(6, 10))
        tb.Label(vac_row, text="Início").pack(side=LEFT)
        tb.Entry(vac_row, textvariable=vac_start_var, width=6).pack(side=LEFT, padx=8)
        tb.Label(vac_row, text="Fim").pack(side=LEFT)
        tb.Entry(vac_row, textvariable=vac_end_var, width=6).pack(side=LEFT, padx=8)
        tb.Label(vac_row, text="(ex: 10 a 20)", bootstyle=SECONDARY).pack(side=LEFT, padx=8)

        def on_save():
            team = team_var.get().strip()
            name = name_var.get().strip()
            role = role_var.get().strip()
            shift = shift_var.get().strip()
            hours = hours_var.get().strip() or "-"

            if not team or not name:
                Messagebox.show_error("Validação", "Preencha TIME e Nome.")
                return

            off_days = self.validate_off_days(safe_int_list_from_csv(off_var.get()), y, m)
            vstart = self.validate_day(vac_start_var.get(), y, m)
            vend = self.validate_day(vac_end_var.get(), y, m)

            if (vstart is None) != (vend is None):
                Messagebox.show_error("Validação", "Para férias, preencha início e fim (ou deixe ambos vazios).")
                return
            if vstart is not None and vend is not None and vstart > vend:
                Messagebox.show_error("Validação", "Férias: início não pode ser maior que fim.")
                return

            if existing:
                existing.team = team
                existing.name = name
                existing.role = role if role in ROLES else "Jr"
                existing.shift = shift if shift in SHIFT_ORDER else SHIFT_ORDER[0]
                existing.hours = hours
                existing.off_days = off_days
                existing.vacation_start = vstart
                existing.vacation_end = vend
            else:
                self.employees.append(
                    Employee(
                        emp_id=new_id(),
                        team=team,
                        name=name,
                        role=role if role in ROLES else "Jr",
                        shift=shift if shift in SHIFT_ORDER else SHIFT_ORDER[0],
                        hours=hours,
                        off_days=off_days,
                        vacation_start=vstart,
                        vacation_end=vend,
                    )
                )

            self.persist()
            self.refresh_employees_table()
            self.refresh_preview_all()
            win.destroy()

        btns = tb.Frame(frame)
        btns.pack(fill=X, pady=16)
        tb.Button(btns, text="Cancelar", bootstyle=SECONDARY, command=win.destroy).pack(side=RIGHT, padx=6)
        tb.Button(btns, text="Salvar", bootstyle=SUCCESS, command=on_save).pack(side=RIGHT, padx=6)

    def refresh_preview_all(self):
        y, m = self.get_year_month()
        days = month_days(y, m)
        headers = ["Time", "Nome", "Cargo", "Horário"] + build_headers_for_month(days)

        for child in self.preview_inner.winfo_children():
            child.destroy()

        for shift in SHIFT_ORDER:
            group = sorted([e for e in self.employees if e.shift == shift], key=lambda x: (x.team.lower(), x.name.lower()))
            lf = tb.Labelframe(self.preview_inner, text=f"Turno: {shift}", padding=8)
            lf.pack(fill=X, expand=True, pady=(0, 10))

            if not group:
                tb.Label(lf, text="(sem colaboradores neste turno)", bootstyle=SECONDARY).pack(anchor=W)
                continue

            data = []
            for e in group:
                off = set(e.off_days or [])
                vac_set = set()
                if isinstance(e.vacation_start, int) and isinstance(e.vacation_end, int) and e.vacation_start <= e.vacation_end:
                    vac_set = set(range(e.vacation_start, e.vacation_end + 1))

                row = [e.team, e.name, e.role, e.hours]
                for d in days:
                    if d.day in vac_set:
                        row.append("Férias")
                    elif d.day in off:
                        row.append("F")
                    else:
                        row.append("P")
                data.append(row)

            cols = [{"text": h, "stretch": True} for h in headers]
            pr_table = Tableview(
                master=lf,
                coldata=cols,
                rowdata=data,
                paginated=False,
                searchable=False,
                autofit=True,
                height=6,
                bootstyle=INFO,
            )
            pr_table.pack(fill=X, expand=True)

        self.preview_canvas.configure(scrollregion=self.preview_canvas.bbox("all"))

    def generate_excel(self):
        y, m = self.get_year_month()
        if not self.employees:
            Messagebox.show_warning("Sem dados", "Cadastre colaboradores antes de gerar.")
            return

        filename = f"Escala_{MONTHS_PT[m]}_{y}.xlsx".replace(" ", "_")
        path = str(Path.cwd() / filename)

        try:
            generate_schedule_xlsx(self.employees, y, m, path)
            self.last_generated_path = path
            self.status_var.set(f"✅ Escala gerada: {path}")
        except Exception as ex:
            Messagebox.show_error("Erro", str(ex))

    def open_folder(self):
        try:
            if self.last_generated_path and os.path.exists(self.last_generated_path):
                os.startfile(os.path.dirname(self.last_generated_path))  # type: ignore[attr-defined]
            else:
                os.startfile(str(Path.cwd()))  # type: ignore[attr-defined]
        except Exception as ex:
            Messagebox.show_error("Erro", str(ex))


if __name__ == "__main__":
    App().mainloop()